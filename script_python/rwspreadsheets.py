import pandas as pd
import sys
import os
import time
import datetime

#from openpyxl import load_workbook
import openpyxl


"""
Ref links:
https://stackoverflow.com/questions/45103927/appending-rows-in-excel-xlswriter
"""

def write_spreadsheet_by_columns():
    # A dict
    # 1st key is the date string, 2nd key is the time string
    tmdict = {"date" : ["2023-05-26", "2023-05-26"],
              "time" : ["12:01:32", "15:58:32"], }

    # A dict
    # each key corresponds to a column
    # The value for the key is the elements of this column
    datadict = {"Beijing" : [0, 1],
            "Guangzhou" : [2, 3],
            "Suzhou" : [4, 5],
            "Hangzhou" : [6, 7],
            "Nanjing" : [8, 9],
            "Xian" : [10, 11],
            "Chengdu" : [12, 13],
            "Chongqing" : [14, 15],
            "Tianjin" : [16, 17],
            "Hefei" : [18, 19],
            "Fuzhou" : [20, 21],
            "Xiamen" : [22, 23],
            "Changsha" : [24, 25],
            "Shanghai" : [26, 27],
            "Shenzhen" : [28, 29],
            "Wuhan" : [30, 31], }

    # Copy 'tmdict' to a new one, and the copy 'datadict'
    # to it as well
    cdict = tmdict.copy()
    cdict.update(datadict)

    # Create a pandas dataframe by this dict data
    #df = pd.DataFrame(datadict)
    df = pd.DataFrame(cdict)

    # Write out a spreadsheet file (.xlsx)
    spfname = 'citynum.xlsx'
    try:
        df.to_excel(spfname, sheet_name='CityResaleNum', index=False)
    except PermissionError as perr:
        #print(perr.what())
        #print(dir(perr))
        print(f"PermissionError: errono = {perr.errno}")
        print(f"PermissionError: strerror = {perr.strerror}")
        print(f"PermissionError: filename = \'{perr.filename}\'")
    except Exception as e:
        print(e.what())


    rdata = pd.read_excel(spfname)
    print(type(rdata))
    print(rdata.head())

def write_spreadsheet_by_rows():

    # City names, each city corresponds to a column in the spreadsheet
    city_list = ["Beijing", "Guangzhou", "Suzhou", "Hangzhou",
                 "Nanjing", "Xian", "Chengdu", "Chongqing",
                 "Tianjin", "Hefei", "Fuzhou", "Xiamen",
                 "Changsha", "Shanghai", "Shenzhen", "Wuhan",]

    # Just example data here: 2 rows, each row is the number for each city
    clen = len(city_list)
    value_2_rows = [[0]*clen, [1]*clen]

    # Inser the 2 rows above into the pandas dataframe first
    df = pd.DataFrame(columns=city_list)
    for i,rval in enumerate(value_2_rows):
        df.loc[i] = rval

    # Write out a spreadsheet file (.xlsx)
    spfname = 'citynum_by_row.xlsx'
    try:
        df.to_excel(spfname, sheet_name='CityResaleNum', index=False)
    except PermissionError as perr:
        #print(perr.what())
        #print(dir(perr))
        print(f"PermissionError: errono = {perr.errno}")
        print(f"PermissionError: strerror = {perr.strerror}")
        print(f"PermissionError: filename = \'{perr.filename}\'")
    except Exception as e:
        print(e.what())

def append_rows_to_existing_spreadsheet():
    spfname = 'citynum_by_row.xlsx'
    if os.path.isfile(spfname) is False:
        write_spreadsheet_by_rows()

    # City names, each city corresponds to a column in the spreadsheet
    city_list = ["Beijing", "Guangzhou", "Suzhou", "Hangzhou",
                 "Nanjing", "Xian", "Chengdu", "Chongqing",
                 "Tianjin", "Hefei", "Fuzhou", "Xiamen",
                 "Changsha", "Shanghai", "Shenzhen", "Wuhan",]

    # More here: extra 3 rows, each row is the number for each city
    clen = len(city_list)
    value_3_rows = [[2]*clen, [3]*clen, [4]*clen]

    # Use openpyxl directly to append rows to an existing spreadsheet
    wb = openpyxl.load_workbook(spfname)
    ws = wb.worksheets[0]

    for rval in value_3_rows:
        ws.append(rval)

    wb.save(spfname)

def create_city_resale_spreadsheet_with_header(spfname, sheet_name="CityResaleNum"):

    if not isinstance(spfname, str):
        return False

    # If file exists, assume this function should not be called
    if os.path.isfile(spfname) is True:
        return False

    dt_cols = ["Date", "Time", "Weekday", "Week"]
    city_cols = [ "Beijing", "Guangzhou", "Suzhou", "Hangzhou",
                  "Nanjing", "Xian", "Chengdu", "Chongqing",
                  "Tianjin", "Hefei", "Fuzhou", "Xiamen",
                  "Changsha", "Shanghai", "Shenzhen", "Wuhan",]

    # Header columns, first 4 are date, time, weekday, week, the rest are city names
    # Current there are 16 cities
    header_columns = dt_cols.copy()
    header_columns.extend(city_cols)

    df = pd.DataFrame(columns=header_columns)
    try:
        df.to_excel(spfname, sheet_name=sheet_name, index=False)
    except PermissionError as perr:
        print(f"PermissionError: errono = {perr.errno}")
        print(f"PermissionError: strerror = {perr.strerror}")
        print(f"PermissionError: filename = \'{perr.filename}\'")
    except Exception as e:
        print(e.what())

    return os.path.isfile(spfname)

def get_date_time_value_list():
    #dtime = datetime.date.fromtimestamp(datetime.datetime.now())
    dtime = datetime.datetime.now()
    # Date string
    date_str = dtime.strftime("%Y-%m-%d")
    # Time string
    time_str = dtime.strftime("%H:%M:%S")
    # Week day for today
    day_str = dtime.strftime("%A")
    # Which week
    week_str = dtime.strftime("%W") # %U also works

    return [date_str, time_str, day_str, week_str]

def append_1_row_to_spreadsheet(spfname, sheet_name="CityResaleNum"):

    if not isinstance(spfname, str):
        return False

    if os.path.isfile(spfname) is False:
        if create_city_resale_spreadsheet_with_header(spfname, sheet_name) is False:
            return False

    #row_val = get_date_time_value_list().extend([13]*16)
    row_val = get_date_time_value_list()
    row_val.extend([13]*16)
    print(row_val)

    # Use openpyxl directly to append rows to an existing spreadsheet
    wb = openpyxl.load_workbook(spfname)
    ws = wb.worksheets[0]
    ws.append(row_val)
    wb.save(spfname)



if __name__ == "__main__":
    #write_spreadsheet_by_columns()
    #write_spreadsheet_by_rows()
    #append_rows_to_existing_spreadsheet()
    append_1_row_to_spreadsheet("city_resale_16.xlsx")


